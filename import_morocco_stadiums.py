import os
import sys
import glob
import re
from openpyxl import load_workbook

# Bootstrap Odoo environment
odoo_path = r"C:\Users\rayan\Documents\dev\odoo_world_cup\odoo"
if odoo_path not in sys.path:
    sys.path.insert(0, odoo_path)

import odoo
from odoo.modules.registry import Registry
from odoo.api import Environment

# Parse capacity from cell
def parse_capacity(val):
    if val is None:
        return 500
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    val_str = str(val).strip()
    digits = re.sub(r'\D', '', val_str)
    if digits:
        return int(digits)
    return 500

# Map state from cell
def parse_state(val):
    if val is None:
        return 'ready'
    val_str = str(val).lower().strip()
    if 'opérationnel' in val_str or 'operationnel' in val_str or 'prêt' in val_str or 'operationel' in val_str:
        return 'ready'
    elif 'construction' in val_str:
        return 'construction'
    elif 'arrêt' in val_str or 'arret' in val_str or 'maintenance' in val_str or 'réhabilitation' in val_str or 'rehabilitation' in val_str:
        return 'maintenance'
    return 'ready'

def main():
    config_path = r"C:\Users\rayan\Documents\dev\odoo_world_cup\odoo\odoo.conf"
    odoo.tools.config.parse_config(['-c', config_path])
    
    db_name = odoo.tools.config.get('db_name') or 'worldcup'
    print(f"Connecting to database: {db_name}")
    
    # Initialize Odoo Registry
    registry = Registry(db_name)
    print("Odoo Registry loaded.")
    
    stade_dir = r"C:\Users\rayan\Documents\dev\odoo_world_cup\stade"
    files = glob.glob(os.path.join(stade_dir, "*.xlsx"))
    print(f"Found {len(files)} Excel files in {stade_dir}")
    
    total_created = 0
    total_skipped = 0
    
    with registry.cursor() as cr:
        env = Environment(cr, odoo.SUPERUSER_ID, {})
        StadiumModel = env['wc.stadium']
        
        for f in files:
            filename = os.path.basename(f)
            if filename.startswith('~$') or filename == '-.xlsx':
                continue
                
            print(f"Processing file: {filename}")
            try:
                wb = load_workbook(f, read_only=True, data_only=True)
                sheet = wb.active
                rows = list(sheet.iter_rows(values_only=True))
                wb.close()
                
                if not rows:
                    print(f"  Empty sheet in {filename}")
                    continue
                    
                header_idx = -1
                for idx, r in enumerate(rows[:15]):
                    if any(isinstance(x, str) and "Nom de l'établissement" in x for x in r if x is not None):
                        header_idx = idx
                        break
                        
                if header_idx == -1:
                    print(f"  Header 'Nom de l'établissement' not found in {filename}")
                    continue
                    
                headers = rows[header_idx]
                
                name_col = -1
                city_col = -1
                prov_col = -1
                cap_col = -1
                addr_col = -1
                state_col = -1
                region_col = -1
                
                for col_idx, h in enumerate(headers):
                    if h is None:
                        continue
                    h_str = str(h).lower()
                    if "nom de l'établissement" in h_str:
                        name_col = col_idx
                    elif "commune" in h_str:
                        city_col = col_idx
                    elif "province" in h_str:
                        prov_col = col_idx
                    elif "capacité" in h_str or "capacite" in h_str:
                        cap_col = col_idx
                    elif "adresse" in h_str:
                        addr_col = col_idx
                    elif "état de l'établissement" in h_str or "etat de l'etablissement" in h_str or "état" in h_str:
                        state_col = col_idx
                    elif "région" in h_str or "region" in h_str:
                        region_col = col_idx
                        
                if name_col == -1:
                    print(f"  Name column not found in headers of {filename}")
                    continue
                    
                file_created = 0
                file_skipped = 0
                for r in rows[header_idx+1:]:
                    if not any(x is not None for x in r):
                        continue
                    name = r[name_col]
                    if name is None or str(name).strip() == '' or str(name).strip() == '_':
                        continue
                        
                    name_str = str(name).strip()
                    city = r[city_col] if city_col != -1 and r[city_col] is not None else (r[prov_col] if prov_col != -1 and r[prov_col] is not None else (r[region_col] if region_col != -1 and r[region_col] is not None else "Morocco"))
                    city_str = str(city).strip()
                    cap = parse_capacity(r[cap_col]) if cap_col != -1 else 500
                    addr = str(r[addr_col]).strip() if addr_col != -1 and r[addr_col] is not None else f"{city_str}, Morocco"
                    state = parse_state(r[state_col]) if state_col != -1 else 'ready'
                    
                    # Check if already exists to avoid duplicates
                    existing = StadiumModel.search([
                        ('name', '=', name_str),
                        ('city', '=', city_str)
                    ])
                    
                    if not existing:
                        StadiumModel.create({
                            'name': name_str,
                            'city': city_str,
                            'capacity': cap,
                            'address': addr,
                            'state': state,
                            'country': 'morocco'
                        })
                        file_created += 1
                    else:
                        file_skipped += 1
                
                print(f"  Done {filename}: Created {file_created}, Skipped {file_skipped}")
                total_created += file_created
                total_skipped += file_skipped
                
            except Exception as e:
                print(f"Error processing {filename}: {e}")
                
        # Commit all database writes
        cr.commit()
        print("Database transaction committed.")
        
    print(f"\nImport Finished!")
    print(f"Total Moroccan Stadiums Created: {total_created}")
    print(f"Total Moroccan Stadiums Skipped: {total_skipped}")

if __name__ == '__main__':
    main()
